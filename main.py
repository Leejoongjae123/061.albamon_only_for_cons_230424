import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow,QMessageBox,QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime,date,timedelta
import numpy
import datetime
from window2 import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import json
import pprint

def get_url_list(pageNo):
    cookies = {
        # '_gid': 'GA1.2.1808740919.1682659603',
        # '_dc_gtm_UA-213826050-1': '1',
        # '_ga': 'GA1.2.588265585.1681308334',
        # 'ConditionId': '5901C861-CB45-4F78-87D8-D25648111214',
        # '_ga_538P897ZYY': 'GS1.1.1682659603.10.1.1682659615.48.0.0',
    }

    headers = {
        # 'authority': 'bff-general.albamon.com',
        'accept': 'application/json',
        # 'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'albamon-domain-type': 'pc',
        'content-type': 'application/json',
        # # 'cookie': '_gid=GA1.2.1808740919.1682659603; _dc_gtm_UA-213826050-1=1; _ga=GA1.2.588265585.1681308334; ConditionId=5901C861-CB45-4F78-87D8-D25648111214; _ga_538P897ZYY=GS1.1.1682659603.10.1.1682659615.48.0.0',
        # 'origin': 'https://www.albamon.com',
        'referer': 'https://www.albamon.com/jobs/part?page=1',
        # 'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
        # 'sec-ch-ua-mobile': '?0',
        # 'sec-ch-ua-platform': '"Windows"',
        # 'sec-fetch-dest': 'empty',
        # 'sec-fetch-mode': 'cors',
        # 'sec-fetch-site': 'same-site',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36',
    }

    json_data = {
        'pagination': {
            'page': int(pageNo),
            'size': 20,
        },
        'recruitListType': 'PART',
        'extensionCondition': {
            'area': {},
            'brand': {},
            'franchise': {},
            'franchiseStore': {},
            'callCenter': {},
            'guerrilla': {},
            'map': {
                'radius': 120,
                'zoom': 0,
            },
            'miniJob': {},
            'ongoing': {},
            'part': {
                'selectedPart': {
                    'categoryCode': '',
                    'code': '',
                },
            },
            'pay': {},
            'preference': {},
            'recent': {},
            'scrap': {},
            'search': {
                'disableExceptedConditions': [],
                'suitBannerNo': 0,
            },
            'season': {},
            'senior': {},
            'shortTerm': {},
            'specUp': {},
            'subway': {},
            'suit': {},
            'teenager': {},
            'town': {
                'similarDongJoin': False,
            },
            'trust': {},
            'university': {},
            'welfare': {},
        },
        'sortTabCondition': {
            'searchPeriodType': 'ALL',
            'sortType': 'DEFAULT',
            'recruitListViewType': 'LIST',
            'latitude': 0,
            'longitude': 0,
        },
        'condition': {
            'areas': [],
            'similarDongJoin': False,
            'parts': [
                {
                    'categoryCode': '8000',
                    'code': '',
                },
            ],
            'workPeriodTypes': [],
            'workWeekTypes': [],
            'workDayTypes': [],
            'workTimeTypes': [],
            'excludeNegoAge': False,
            'employmentTypes': [],
        },
    }

    response = requests.post('https://bff-general.albamon.com/recruit/search', cookies=cookies, headers=headers,
                             json=json_data)
    print(response.text)
    result_raw=json.loads(response.text)
    results=result_raw['base']['normal']['collection']
    url_list=[]
    for index,result in enumerate(results):
        # print(result)
        name=result['recruitTitle']
        recruitNo=result['recruitNo']
        print(index+1,"번째 ",'name:',name,"recruitNo:",recruitNo)
        url='https://www.albamon.com/jobs/detail/{}'.format(recruitNo)
        url_list.append([pageNo,url])
    return url_list

class Thread(QThread,QMainWindow,Ui_MainWindow):
    # 초기화 메서드 구현

    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성
    user_signal2 = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,list_flag,time_delay,page_limit,page_start,fname):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.list_flag=list_flag
        self.time_delay=time_delay
        self.page_limit=page_limit
        self.page_start=page_start
        self.login_flag=False
        self.fname=fname
    def run(self):
        if self.list_flag == False:
            print("리스트가져오기시작")
            wb = openpyxl.Workbook()
            ws = wb.active
            data = ['페이지', 'URL', '상호', '연락처', '주소']
            ws.append(data)

            for i in range(1, 9999999):
                if i>self.page_limit: # 1번째 페이지만 하기 위한 테스트 코드
                    print("{}번째 페이지 까지만 크롤링..".format(i-1))
                    break
                if i<self.page_start:
                    print("해당 없어서 페이지 스킵")
                    continue

                # self.parent.textEdit.append("{}번째 페이지의 리스트 가져오는 중...".format(i))
                text="{}번째 페이지의 리스트 가져오는 중...".format(i)
                self.user_signal.emit(text)
                time_now = datetime.datetime.now()
                time_now = time_now.strftime("%H%M%S")
                print("{}번째 페이지 링크 가져오는 중...{}".format(i, time_now))
                url_list=get_url_list(i)

                if len(url_list) >= 1:
                    print("페이지있음")
                else:
                    print("페이지없음")
                    break
                for index, url_elem in enumerate(url_list):
                    data=[url_elem[0],url_elem[1]]
                    ws.append(data)
                time.sleep(random.randint(10, 20) * 0.1)
            time_now=datetime.datetime.now()
            time_now_string=time_now.strftime("%Y%m%d_%H%M%S")
            wb.save('result_{}.xlsx'.format(time_now_string))
            text="URL 리스트 모으기가 완료되었습니다."
            self.user_signal2.emit(text)


        if self.list_flag == True:
            print("상세정보검색시작")
            wb = openpyxl.load_workbook(self.fname)
            ws = wb.active
            no_row = ws.max_row
            info_list = []

            for i in range(2, no_row + 1):
                empty = True
                page_num = ws.cell(row=i, column=1).value
                url = ws.cell(row=i, column=2).value
                info = ws.cell(row=i, column=3).value
                if page_num == "" or page_num == None:
                    print("빈행임")
                    break
                if info != None:
                    empty = False
                info = [page_num, url, empty]
                info_list.append(info)

            print(info_list)


            print('info_list:',info_list)


            for index, info in enumerate(info_list):
                empty = info[2]
                if empty == False:
                    continue
                page_num = info[0]
                url = info[1]

                print("{}번째 크롤링중...{}".format(index + 1,url))

                # self.parent.textEdit.append("{}번째 행 크롤링 중...".format(index + 1))
                text="{}번째 행 크롤링 중...".format(index + 1)
                self.user_signal.emit(text)

                #-------------------클라이언트 계정

                # 알바몬에가서 SNS이름으로 된 거 찾아서 직접 로그인 해보고 로그인 정보 가져와서 requests 날려야 한다.!

                cookies = {
                    'ConditionId': '5785EB64-139D-4647-80E2-062B1103755E',
                    '_gid': 'GA1.2.109883684.1683012339',
                    '_gac_UA-213826050-1': '1.1683012345.Cj0KCQjw6cKiBhD5ARIsAKXUdyaqCPh5EmNwSjuZX_k3QyhmYGfA72GuPxbQZYVfLISHFXneQgvhTUEaAlPHEALw_wcB',
                    '_dc_gtm_UA-213826050-1': '1',
                    '_ga': 'GA1.2.588265585.1681308334',
                    '_ga_538P897ZYY': 'GS1.1.1683012339.14.1.1683015252.38.0.0',
                }

                headers = {
                    'authority': 'bff-general.albamon.com',
                    'accept': 'application/json',
                    'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                    'albamon-domain-type': 'pc',
                    'content-type': 'application/json',
                    # 'cookie': 'ConditionId=5785EB64-139D-4647-80E2-062B1103755E; _gid=GA1.2.109883684.1683012339; _gac_UA-213826050-1=1.1683012345.Cj0KCQjw6cKiBhD5ARIsAKXUdyaqCPh5EmNwSjuZX_k3QyhmYGfA72GuPxbQZYVfLISHFXneQgvhTUEaAlPHEALw_wcB; _dc_gtm_UA-213826050-1=1; _ga=GA1.2.588265585.1681308334; _ga_538P897ZYY=GS1.1.1683012339.14.1.1683015252.38.0.0',
                    'origin': 'https://www.albamon.com',
                    'referer': 'https://www.albamon.com/user-account/login?linkType=logout&memberType=PERSONAL',
                    'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
                    'sec-ch-ua-mobile': '?0',
                    'sec-ch-ua-platform': '"Windows"',
                    'sec-fetch-dest': 'empty',
                    'sec-fetch-mode': 'cors',
                    'sec-fetch-site': 'same-site',
                    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36',
                }

                json_data = {
                    'snsType': 'GOOGLE',
                    'snsId': '106928325900762547836',
                    'snsEmail': 'tabeoringim@gmail.com',
                    'userName': 'gim tan',
                    'snsCode': 'ya29.a0AWY7CknPL5O4BQcXv55s1HkViul0YSdR8cczj8HwVafw3a38zed2fUtqdOmq2w4-qPHe1VJeOAKvkeQ_mpIe-qtXHj-C4wQrDW1hyB2amyzbPXq7k-niD4I4qXmDjgTaUHA_kAS2jDXjHwlsgRxgbKkT631GaCgYKAZcSARMSFQG1tDrpi_NrZIScb0Mn20wKzRxiew0163',
                }
                #-------------------클라이언트 계정


                session = requests.session()

                response = session.post('https://bff-general.albamon.com/member/login/sns', cookies=cookies,
                                        headers=headers, json=json_data)

                print("status_code:", response.status_code)
                count = 0


                time_now = datetime.datetime.now()
                time_now_string = time_now.strftime("%H%M%S")
                response = session.get(url)
                soup = BeautifulSoup(response.text, 'lxml')
                result_raw = str(soup.find('script', attrs={'id': '__NEXT_DATA__'}))
                position_fr = result_raw.find("{")
                position_rr = result_raw.rfind("}")
                # print(position_fr,position_rr)
                result = result_raw[position_fr:position_rr + 1]
                # print(result)
                data = json.loads(result)
                # pprint.pprint(data)
                try:
                    phone_number=data['props']['pageProps']['data']['viewData']['phoneNumber'][0]['phoneNumber']
                except:
                    phone_number=""
                # regex=re.compile("010-")fff
                # result=regex.findall(phone_number)
                # if len(result)==0:
                #     phone_number="핸드폰 번호 없음"
                print('phone_number:',phone_number)
                # pprint.pprint(data['props']['pageProps']['data']['viewData']['phoneNumber'][0]['phoneNumber'])
                try:
                    company_name=data['props']['pageProps']['data']['companyData']['companyName']
                except:
                    company_name=""
                print('company_name:',company_name)
                try:
                    company_address=data['props']['pageProps']['data']['companyData']['address']
                except:
                    company_address=""
                print('company_address:',company_address)

                ws.cell(row=index+2,column=3).value=company_name
                ws.cell(row=index+2, column=4).value = phone_number
                ws.cell(row=index+2,column=5).value=company_address
                wb.save(self.fname)
                time.sleep(self.time_delay)
                print("-------------------------------------------------------")
            # self.parent.textEdit.append("작업이 완료 되었습니다")
            text="작업이 완료 되었습니다"
            self.user_signal.emit(text)
            text="상세 데이타 모으기가 완료되었습니다."
            self.user_signal2.emit(text)

class Example(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path="C:"
        self.index=None
        self.setupUi(self)
        self.setSlot()
        self.show()
        self.login_flag=False
        QApplication.processEvents()
        self.fname=""
        self.lineEdit_2.setText("1")


    def start(self):
        self.list_flag = False
        self.time_delay=int(self.lineEdit_2.text())
        print("딜레이는:",self.time_delay)
        self.page_start = int(self.lineEdit_4.text())
        print("페이지스타트는:", self.page_start)
        self.page_limit=self.lineEdit_3.text()
        if self.page_limit=="" or self.page_limit==None :
            self.page_limit=999999
        else:
            self.page_limit=int(self.page_limit)
        print("페이지리밋은:",self.page_limit)
        print("리스트플래그:",self.list_flag)

        self.x=Thread(self,self.list_flag,self.time_delay,self.page_limit,self.page_start,self.fname)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def start_again(self):
        self.list_flag = True
        self.time_delay=int(self.lineEdit_2.text())
        print("딜레이는:",self.time_delay)
        self.page_start = int(self.lineEdit_4.text())
        print("페이지스타트는:", self.page_start)
        self.page_limit=self.lineEdit_3.text()
        if self.page_limit=="" or self.page_limit==None :
            self.page_limit=999999
        else:
            self.page_limit=int(self.page_limit)
        print("페이지리밋은:",self.page_limit)
        print("리스트플래그:",self.list_flag)

        self.x=Thread(self,self.list_flag,self.time_delay,self.page_limit,self.page_start,self.fname)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.user_signal2.connect(self.slot2)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self,data):
        self.textEdit.append(str(data))

    def slot2(self,data2):
        QMessageBox.information(self, "완료창",data2)

    def search(self):
        print("find")
        self.fname = QFileDialog.getOpenFileName(self, "Open file", './')[0]
        print(self.fname)
        self.lineEdit_5.setText(self.fname)

    def setSlot(self):
        pass
    def setIndex(self,index):
        pass
    def login(self):
        print('로그인시도')
        self.x.login()
    def quit(self):
        QCoreApplication.instance().quit()
    def complete(self):
        QMessageBox.information(self, "완료창", "작업이 완료 되었습니다.")

app=QApplication([])
ex=Example()
sys.exit(app.exec_())



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())







